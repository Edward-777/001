"""Downloadable report files (xlsx). The same report data the chat summarizes is
also offered as real spreadsheets the user can hand to their accountant.

`closing_package` is the full month-end binder (every sheet); the others are
single focused reports.
"""
from __future__ import annotations

import io
from datetime import date

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from . import reports
from .ledger_models import JournalEntry, JournalLine
from .models import Account

REPORT_KINDS = (
    "closing_package", "financials", "cash_flow", "trial_balance",
    "general_ledger", "journal_entries", "ap_aging", "ar_aging", "inventory",
)


def latest_active_period(session: Session) -> str:
    """The YYYY-MM of the most recent journal activity (so 'current' financials
    land on real data, not an empty future month). Falls back to this month."""
    d = session.scalar(
        select(func.max(JournalEntry.entry_date)).where(
            JournalEntry.status.in_(["posted", "reversed"]))
    )
    d = d or date.today()
    return f"{d.year:04d}-{d.month:02d}"


def _h(ws, values) -> None:
    """Append a row and bold it (titles, headers, totals)."""
    from openpyxl.styles import Font

    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


# ---- one builder per sheet (reused across report kinds) -----------------

def _income_statement(wb, session, period, start, end) -> None:
    is_ = reports.income_statement(session, start=start, end=end)
    s = wb.create_sheet("Income Statement")
    _h(s, [f"Income Statement — {period}"])
    _h(s, ["Account", "Amount"])
    for r in is_["revenue"]:
        s.append([f"Revenue: {r['code']} {r['name']}", float(r["balance"])])
    for r in is_["expenses"]:
        s.append([f"Expense: {r['code']} {r['name']}", float(r["balance"])])
    _h(s, ["Net income", float(is_["net_income"])])


def _balance_sheet(wb, session, as_of) -> None:
    bs = reports.balance_sheet(session, as_of=as_of)
    b = wb.create_sheet("Balance Sheet")
    _h(b, [f"Balance Sheet — as of {as_of}"])
    _h(b, ["Account", "Amount"])
    for grp, label in (("assets", "Asset"), ("liabilities", "Liability"), ("equity", "Equity")):
        for r in bs[grp]:
            b.append([f"{label}: {r['code']} {r['name']}", float(r["balance"])])
    b.append(["Total assets", float(bs["total_assets"])])
    b.append(["Total liabilities", float(bs["total_liabilities"])])
    _h(b, ["Total equity", float(bs["total_equity"])])


def _cash_flow(wb, session, period, start, end) -> None:
    cf = reports.cash_flow(session, start=start, end=end)
    s = wb.create_sheet("Cash Flow")
    _h(s, [f"Statement of Cash Flows — {period} ({cf['method']})"])
    s.append(["Beginning cash", float(cf["beginning_cash"])])
    s.append(["Operating", float(cf["operating"])])
    s.append(["Investing/Financing", float(cf["investing_financing"])])
    s.append(["Net change", float(cf["net_change"])])
    _h(s, ["Ending cash", float(cf["ending_cash"])])


def _trial_balance(wb, session, as_of) -> None:
    tb = reports.trial_balance(session, as_of=as_of)
    s = wb.create_sheet("Trial Balance")
    _h(s, [f"Trial Balance — as of {as_of}"])
    _h(s, ["Code", "Account", "Type", "Debit", "Credit"])
    for r in tb["rows"]:
        s.append([r["code"], r["name"], r["type"], float(r["debit"]), float(r["credit"])])
    _h(s, ["", "TOTAL", "", float(tb["total_debit"]), float(tb["total_credit"])])


def _general_ledger(wb, session, start, end) -> None:
    s = wb.create_sheet("General Ledger")
    _h(s, [f"General Ledger — {start or 'inception'} to {end}"])
    accts = sorted(reports._accounts(session).values(), key=lambda a: a.code)
    for a in accts:
        rows = reports.general_ledger(session, a.id, start=start, end=end)
        if not rows:
            continue
        _h(s, [f"{a.code} {a.name}"])
        _h(s, ["JE", "Date", "Description", "Debit", "Credit"])
        for r in rows:
            s.append([r["je_no"], str(r["date"]), r["description"],
                      float(r["debit"]), float(r["credit"])])


def _journal_entries(wb, session, start, end) -> None:
    s = wb.create_sheet("Journal Entries")
    _h(s, [f"Journal Entries — {start or 'inception'} to {end}"])
    _h(s, ["JE", "Date", "Source", "Status", "Account", "Debit", "Credit"])
    stmt = (
        select(JournalEntry.je_no, JournalEntry.entry_date, JournalEntry.source_type,
               JournalEntry.status, Account.code, Account.name,
               JournalLine.debit, JournalLine.credit)
        .join(JournalLine, JournalLine.je_id == JournalEntry.id)
        .join(Account, Account.id == JournalLine.account_id)
        .where(JournalEntry.entry_date <= end, JournalEntry.status.in_(["posted", "reversed"]))
        .order_by(JournalEntry.entry_date, JournalEntry.id, JournalLine.id)
    )
    if start is not None:
        stmt = stmt.where(JournalEntry.entry_date >= start)
    for je_no, d, src, st, code, name, dr, cr in session.execute(stmt).all():
        s.append([je_no, str(d), src, st, f"{code} {name}", float(dr), float(cr)])


def _aging(wb, session, kind, as_of) -> None:
    ag = (reports.ap_aging if kind == "ap_aging" else reports.ar_aging)(session, as_of=as_of)
    s = wb.create_sheet("AP Aging" if kind == "ap_aging" else "AR Aging")
    _h(s, [f"{'AP' if kind == 'ap_aging' else 'AR'} Aging — as of {as_of}"])
    _h(s, ["Document", "Party id", "Due date", "Balance", "Bucket"])
    for r in ag["rows"]:
        doc = r.get("bill_no") or r.get("invoice_no")
        party = r.get("vendor_id") or r.get("customer_id")
        s.append([doc, party, str(r.get("due_date") or ""), float(r["balance"]), r["bucket"]])
    _h(s, ["", "", "TOTAL", float(ag["total"]), ""])


def _inventory(wb, session) -> None:
    val = reports.inventory_valuation(session)
    s = wb.create_sheet("Inventory")
    _h(s, ["Inventory Valuation"])
    _h(s, ["Product id", "Qty on hand", "Avg cost", "Value"])
    for r in val["rows"]:
        s.append([r["product_id"], float(r["qty"]), float(r["avg_unit_cost"]), float(r["value"])])
    _h(s, ["", "", "TOTAL", float(val["total_value"])])


def build_report_xlsx(session: Session, kind: str, period: str | None = None) -> tuple[str, bytes]:
    """Return (filename, xlsx-bytes) for a report kind. 'closing_package' bundles
    every statement + supporting schedule into one workbook."""
    from openpyxl import Workbook  # lazy: only needed when a report is exported

    period = period or latest_active_period(session)
    start, end = reports._month_bounds(period)
    wb = Workbook()
    wb.remove(wb.active)

    if kind == "closing_package":
        _balance_sheet(wb, session, end)
        _income_statement(wb, session, period, start, end)
        _cash_flow(wb, session, period, start, end)
        _trial_balance(wb, session, end)
        _general_ledger(wb, session, None, end)   # full ledger through period end
        _journal_entries(wb, session, None, end)  # all entries through period end
        _aging(wb, session, "ap_aging", end)
        _aging(wb, session, "ar_aging", end)
        _inventory(wb, session)
    elif kind == "financials":
        _balance_sheet(wb, session, end)
        _income_statement(wb, session, period, start, end)
        _cash_flow(wb, session, period, start, end)
    elif kind == "cash_flow":
        _cash_flow(wb, session, period, start, end)
    elif kind == "trial_balance":
        _trial_balance(wb, session, end)
    elif kind == "general_ledger":
        _general_ledger(wb, session, start, end)
    elif kind == "journal_entries":
        _journal_entries(wb, session, start, end)
    elif kind in ("ap_aging", "ar_aging"):
        _aging(wb, session, kind, end)
    elif kind == "inventory":
        _inventory(wb, session)
    else:
        raise ValueError(f"unknown report kind: {kind}")

    if not wb.sheetnames:  # never hand back an empty workbook
        wb.create_sheet("Empty").append(["No data for this report/period."])
    buf = io.BytesIO()
    wb.save(buf)
    return f"{kind}_{period}.xlsx", buf.getvalue()
