"""Downloadable purchase-order document (xlsx) — what the user sends to the vendor
until email delivery ships. Mirrors sales.documents: builder returns (filename, bytes).
"""
from __future__ import annotations

import io

from sqlalchemy.orm import Session

from . import service as proc


def _bold(ws, values) -> None:
    from openpyxl.styles import Font

    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def build_po_xlsx(session: Session, po_id: int) -> tuple[str, bytes]:
    from openpyxl import Workbook

    po = proc.get_po(session, po_id)
    if po is None:
        raise ValueError("PO not found")
    vendor = proc.get_vendor(session, po.vendor_id) if po.vendor_id else None
    wb = Workbook()
    s = wb.active
    s.title = "Purchase Order"
    _bold(s, ["PURCHASE ORDER"])
    s.append(["PO no", po.po_no])
    s.append(["Order date", str(po.order_date or "")])
    s.append(["Expected date", str(po.expected_date or "")])
    s.append(["Vendor", vendor.name if vendor else ""])
    if vendor and vendor.address:
        s.append(["Vendor address", vendor.address])
    if vendor and vendor.email:
        s.append(["Vendor email", vendor.email])
    s.append([])
    _bold(s, ["Description", "Qty", "Unit price", "Amount"])
    for ln in po.lines:
        s.append([ln.description, float(ln.qty_ordered), float(ln.unit_price),
                  float(ln.amount)])
    s.append([])
    s.append(["", "", "Subtotal", float(po.subtotal)])
    s.append(["", "", "Tax", float(po.tax)])
    _bold(s, ["", "", "Total", float(po.total)])
    buf = io.BytesIO()
    wb.save(buf)
    return f"PO_{po.po_no}.xlsx", buf.getvalue()
