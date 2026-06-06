"""Downloadable O2C documents (xlsx): quote/estimate, packing list, invoice.

Same idea as accounting.export — the data the pipeline tracks is also handed to
the customer as real documents. Each builder returns (filename, xlsx-bytes).
"""
from __future__ import annotations

import io

from sqlalchemy.orm import Session

from . import fulfillment as ful
from . import service as sales
from .models import ARInvoice


def _bold(ws, values) -> None:
    from openpyxl.styles import Font

    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def _customer_name(session: Session, customer_id: int) -> str:
    c = sales.get_customer(session, customer_id)
    return c.name if c else f"Customer #{customer_id}"


def build_quote_xlsx(session: Session, quote_id: int) -> tuple[str, bytes]:
    from openpyxl import Workbook

    q = ful.get_quote(session, quote_id)
    if q is None:
        raise ValueError("quote not found")
    wb = Workbook()
    s = wb.active
    s.title = "Quote"
    _bold(s, ["QUOTE / ESTIMATE"])
    s.append(["Quote no", q.quote_no])
    s.append(["Date", str(q.quote_date or "")])
    s.append(["Valid until", str(q.valid_until or "")])
    s.append(["Customer", _customer_name(session, q.customer_id)])
    s.append([])
    _bold(s, ["Description", "Qty", "Unit price", "Amount"])
    for ln in q.lines:
        s.append([ln.description, float(ln.qty), float(ln.unit_price), float(ln.amount)])
    s.append([])
    s.append(["", "", "Subtotal", float(q.subtotal)])
    s.append(["", "", "Tax", float(q.tax_amount)])
    _bold(s, ["", "", "Total", float(q.total)])
    buf = io.BytesIO()
    wb.save(buf)
    return f"Quote_{q.quote_no}.xlsx", buf.getvalue()


def build_packing_list_xlsx(session: Session, shipment_id: int) -> tuple[str, bytes]:
    from openpyxl import Workbook

    sh = ful.get_shipment(session, shipment_id)
    if sh is None:
        raise ValueError("shipment not found")
    wb = Workbook()
    s = wb.active
    s.title = "Packing List"
    _bold(s, ["PACKING LIST"])
    s.append(["Shipment no", sh.shipment_no])
    s.append(["Ship date", str(sh.ship_date or "")])
    s.append(["Carrier", sh.carrier or ""])
    s.append(["Tracking no", sh.tracking_no or ""])
    s.append(["Customer", _customer_name(session, sh.customer_id)])
    s.append([])
    _bold(s, ["Description", "Qty shipped"])   # no prices on a packing list
    for ln in sh.lines:
        s.append([ln.description, float(ln.qty)])
    buf = io.BytesIO()
    wb.save(buf)
    return f"PackingList_{sh.shipment_no}.xlsx", buf.getvalue()


def build_invoice_xlsx(session: Session, invoice_id: int) -> tuple[str, bytes]:
    from openpyxl import Workbook

    inv = session.get(ARInvoice, invoice_id)
    if inv is None:
        raise ValueError("invoice not found")
    # the originating quote (if any) carries the customer PO
    po = ""
    if inv.so_id is not None:
        q = ful._quote_for_so(session, inv.so_id)
        po = q.customer_po if q else ""
    wb = Workbook()
    s = wb.active
    s.title = "Invoice"
    _bold(s, ["INVOICE"])
    s.append(["Invoice no", inv.invoice_no])
    s.append(["Date", str(inv.invoice_date or "")])
    s.append(["Due date", str(inv.due_date or "")])
    s.append(["Customer", _customer_name(session, inv.customer_id)])
    s.append(["Customer PO", po])
    s.append([])
    _bold(s, ["Description", "Qty", "Unit price", "Amount"])
    for ln in inv.lines:
        s.append([ln.description, float(ln.qty), float(ln.unit_price), float(ln.amount)])
    s.append([])
    s.append(["", "", "Subtotal", float(inv.subtotal)])
    s.append(["", "", "Tax", float(inv.tax_amount)])
    _bold(s, ["", "", "Total", float(inv.total)])
    s.append(["", "", "Balance due", float(inv.balance)])
    buf = io.BytesIO()
    wb.save(buf)
    return f"Invoice_{inv.invoice_no}.xlsx", buf.getvalue()
