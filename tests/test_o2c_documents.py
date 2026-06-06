"""O2C document generation: quote / packing list / invoice xlsx."""
import io

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.core.db import Base
from app.modules import (  # noqa: F401  register all tables
    accounting, approval, assets, auth, bank, documents, expense,
    fleet, hr, inventory, notifications, procurement, sales,
)
from app.modules.sales import documents as docs
from app.modules.sales import fulfillment as ful
from app.modules.sales import service as sls


@pytest.fixture
def session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


def _rows(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    s = wb.active
    return [[c.value for c in row] for row in s.iter_rows()]


def _pipeline(session):
    c = sls.create_customer(session, name="BigCo Inc")
    quote = ful.create_quote(session, customer_id=c.id, lines=[
        {"description": "Widget Pro", "qty": 10, "unit_price": 1500}])
    so = ful.accept_quote(session, quote.id, customer_po="PO-99821")
    sh = ful.create_shipment(session, so_id=so.id, carrier="FedEx", tracking_no="FX1")
    invoice = ful.invoice_order(session, so.id)
    return quote, sh, invoice


def test_quote_document(session):
    quote, _, _ = _pipeline(session)
    fname, data = docs.build_quote_xlsx(session, quote.id)
    assert fname == f"Quote_{quote.quote_no}.xlsx"
    flat = [str(c) for row in _rows(data) for c in row if c is not None]
    assert "QUOTE / ESTIMATE" in flat
    assert "Widget Pro" in flat and "BigCo Inc" in flat
    assert "15000.0" in flat or 15000.0 in [c for row in _rows(data) for c in row]


def test_packing_list_document(session):
    _, sh, _ = _pipeline(session)
    fname, data = docs.build_packing_list_xlsx(session, sh.id)
    assert fname == f"PackingList_{sh.shipment_no}.xlsx"
    flat = [str(c) for row in _rows(data) for c in row if c is not None]
    assert "PACKING LIST" in flat
    assert "FedEx" in flat and "FX1" in flat and "Widget Pro" in flat


def test_invoice_document_carries_po(session):
    _, _, invoice = _pipeline(session)
    fname, data = docs.build_invoice_xlsx(session, invoice.id)
    assert fname == f"Invoice_{invoice.invoice_no}.xlsx"
    flat = [str(c) for row in _rows(data) for c in row if c is not None]
    assert "INVOICE" in flat
    assert "PO-99821" in flat   # customer PO flows onto the invoice
    assert "Balance due" in flat
